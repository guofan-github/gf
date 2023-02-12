from openpyxl import load_workbook


class read_excel():

    def get_data(self, excel_name, sheet_name):

        # 打开表格
        wd = load_workbook(f"../data/{excel_name}.xlsx")

        # 指定sheet
        ws = wd[sheet_name]

        # 读取出来的格式[["",""],[]]
        rows = ws.max_row
        cols = ws.max_column

        info_list = []
        for r in range(2, rows+1):
            list = []
            for c in range(1, cols+1):
                list.append(ws.cell(r,c).value)

            info_list.append(list)
        return info_list


if __name__ == '__main__':
    a = read_excel().get_data("jcss修改", '修改')
    print(a)




