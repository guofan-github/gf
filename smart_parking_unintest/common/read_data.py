import os

from openpyxl import load_workbook


class Read:

    def get_data(self, filename, sheet):
        # 打开表格
        dir_path = os.path.dirname(os.path.dirname(__file__))
        wb = load_workbook(f'{dir_path}/data/{filename}.xlsx')
        # 指定sheet
        ws = wb[f'{sheet}']
        # 最大行、列
        rows = ws.max_row
        cols = ws.max_column

        data_list = []
        # 循环拿到表格中数据
        for r in range(2, rows + 1):
            # 定义个空列表，来保存获取的信息
            lines = []
            for c in range(1, cols + 1):
                lines.append(ws.cell(r, c).value)
            data_list.append(lines)
        wb.close()
        return data_list



if __name__ == '__main__':
    a= Read().get_data('API测试数据', "api")


